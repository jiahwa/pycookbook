#!/usr/bin/env python3
# install pywin32 xlrd openpyxl xlutils
from win32com.client import Dispatch
import openpyxl
import xlrd
from xlutils.copy import copy

def getGoal(name, goals):
    name, sep, tail = name.partition('_')

    for item in goals:
        _item, sep, tail = item[0].partition('_')
        if name == _item:
            return item[1]

# define constant
AREA_DICT = {}

# AREA_DICT["demo"] = (('read_path', 'read_sheet', ('read_row_begin', 'read_row_end'), ('read_col_select', 'read_col_copy')),
#                      ('write_path', 'write_sheet', ('write_row_begin', 'write_row_end'), ('write_col_select', 'write_col_copy'))


# times 0
# AREA_DICT["nanjing"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\2021应届毕业生前端培训阅卷.xlsx', '3.考核pc', (116, 179), (1, 23)),
#                         ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\南京应届生名单-阅卷成绩.xls', '南京分配-成绩', (1, 62), (4, 5)))

# AREA_DICT["shenzhen"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\2021应届毕业生前端培训阅卷.xlsx', '3.考核pc', (242, 289), (1, 23)),
#                         ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\2021届深圳花名册-阅卷成绩.xlsx', 'Sheet3', (1, 47), (1, 11)))

# AREA_DICT["beijing"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\2021应届毕业生前端培训阅卷.xlsx', '3.考核pc', (4, 116), (1, 23)),
#                         ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\北京地区应届生PC实战-阅卷成绩.xlsx', '技术人员', (1, 113), (2, 3)))

# AREA_DICT["shanghai"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\2021应届毕业生前端培训阅卷.xlsx', '3.考核pc', (204, 239), (1, 23)),
#                         ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\上海培训人员名单-阅卷成绩.xlsx', 'Sheet1', (1, 35), (2, 3)))

# AREA_DICT["suzhou"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\2021应届毕业生前端培训阅卷.xlsx', '3.考核pc', (182, 203), (1, 23)),
#                         ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\苏州2021年应届生名单-前端阅卷成绩.xls', 'Sheet1', (1, 22), (1, 2)))

# AREA_DICT["chengdu"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\2021应届毕业生前端培训阅卷.xlsx', '3.考核pc', (292, 326), (1, 23)),
#                         ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\成都现场培训名单（更正）-阅卷成绩.xlsx', 'Sheet1', (1, 34), (2, 8)))

# times 1
# AREA_DICT["nanjing"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\考试成绩20210803-09_19_05.xlsx', 'Sheet1', (1, 310), (2, 4)),
#                         ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\2021应届毕业生前端培训阅卷-北京.xlsx', '南京', (1, 6), (0, 5)))

# AREA_DICT["beijing"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\考试成绩20210803-09_19_05.xlsx', 'Sheet1', (1, 310), (2, 4)),
#                         ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\2021应届毕业生前端培训阅卷-北京.xlsx', '北京', (1, 112), (0, 5)))

# times 2
AREA_DICT["nanjing"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\理论考试成绩20210803-09_19_05.xlsx', 'Sheet1', (1, 310), (2, 4)),
                        ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\南京应届生名单-阅卷成绩.xls', '南京分配-成绩', (1, 62), (4, 6)))

AREA_DICT["shenzhen"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\理论考试成绩20210803-09_19_05.xlsx', 'Sheet1', (1, 310), (2, 4)),
                        ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\2021届深圳花名册-阅卷成绩.xlsx', 'Sheet3', (1, 47), (1, 12)))

AREA_DICT["beijing"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\理论考试成绩20210803-09_19_05.xlsx', 'Sheet1', (1, 310), (2, 4)),
                        ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\北京地区应届生PC实战-阅卷成绩.xlsx', '技术人员', (1, 113), (2, 4)))

AREA_DICT["shanghai"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\理论考试成绩20210803-09_19_05.xlsx', 'Sheet1', (1, 310), (2, 4)),
                        ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\上海培训人员名单-阅卷成绩.xlsx', 'Sheet1', (1, 35), (2, 4)))

AREA_DICT["suzhou"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\理论考试成绩20210803-09_19_05.xlsx', 'Sheet1', (1, 310), (2, 4)),
                        ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\苏州2021年应届生名单-前端阅卷成绩.xls', 'Sheet1', (1, 22), (1, 3)))

AREA_DICT["chengdu"] = (('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\理论考试成绩20210803-09_19_05.xlsx', 'Sheet1', (1, 310), (2, 4)),
                        ('C:\\Users\\Zz\\Desktop\\python处理考试成绩\\成都现场培训名单（更正）-阅卷成绩.xlsx', 'Sheet1', (1, 34), (2, 9)))

def main(area):
    print(area)
    read_file = AREA_DICT[area][0][0]
    read_sheet = AREA_DICT[area][0][1]

    write_file = AREA_DICT[area][1][0]
    write_sheet = AREA_DICT[area][1][1]

    # initial open and close
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(read_file)
    xlBook.Save()
    xlBook.Close()

    # read only
    workbook_read = openpyxl.load_workbook(read_file, data_only=True)
    sheet_read = workbook_read.get_sheet_by_name(read_sheet)

    # define save goals
    goals = []

    read_row_begin = AREA_DICT[area][0][2][0]
    read_row_end = AREA_DICT[area][0][2][1]
    read_col_select = AREA_DICT[area][0][3][0]
    read_col_copy = AREA_DICT[area][0][3][1]

    for row in range(read_row_begin, read_row_end):
        col1 = sheet_read.cell(row=row, column=read_col_select).value
        col23 = sheet_read.cell(row=row, column=read_col_copy).value
        # print(sheet_read.cell(row=row, column=1).value, , end=' ')
        # print('\r')
        goals.append((col1,col23))
    print(goals)

    # write only
    # workbook_write = xlrd.open_workbook(write_file, formatting_info=True)
    workbook_write = xlrd.open_workbook(write_file) # xlsx要注释掉formatting_info=True, reason: raise NotImplementedError("formatting_info=True not yet implemented")
    sheet_write = workbook_write.sheet_by_name(write_sheet)
    # nrows = sheet_write.nrows

    # copy copyer
    copy_xlsx = copy(workbook_write)
    target_sheet = copy_xlsx.get_sheet(0)
    # target_sheet = copy_xlsx.get_sheet_by_name

    write_row_begin = AREA_DICT[area][1][2][0]
    write_row_end = AREA_DICT[area][1][2][1]
    write_col_select = AREA_DICT[area][1][3][0]
    write_col_copy = AREA_DICT[area][1][3][1]

    for row in range(write_row_begin, write_row_end):
        col4 = sheet_write.cell(row, write_col_select).value
        goal = getGoal(col4, goals)
        if goal is not None:
            # print(row, write_col_copy, goal)
            target_sheet.write(row, write_col_copy, goal)

    # save
    copy_xlsx.save(write_file)

# can be modified here
# area = 'nanjing'
# area = 'shenzhen'
# area = 'beijing'
# area = 'shanghai'
# area = 'suzhou'
# area = 'chengdu'

# execute six areas
for i in AREA_DICT:
    main(i)